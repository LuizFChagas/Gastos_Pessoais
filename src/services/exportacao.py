import csv
import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.graphics.shapes import Drawing
from reportlab.graphics.charts.piecharts import Pie

FORMA_PAGAMENTO_LABEL = {
    "debito": "Débito", "credito": "Crédito", "pix": "Pix",
    "dinheiro": "Dinheiro", "outro": "Outro",
}

CORES_GRAFICO = [
    "#10b981", "#3b82f6", "#f97316", "#ec4899", "#8b5cf6", "#facc15",
    "#ef4444", "#14b8a6", "#6366f1", "#0ea5e9", "#78716c", "#06b6d4",
]


def _linha(g) -> dict:
    return {
        "data": g.data_hora.strftime("%d/%m/%Y") if g.data_hora else "",
        "descricao": g.descricao or "",
        "categoria": (g.categoria or "outros").capitalize(),
        "tipo": "Entrada" if g.tipo == "entrada" else "Saída",
        "forma_pagamento": FORMA_PAGAMENTO_LABEL.get(g.forma_pagamento, g.forma_pagamento or "—"),
        "banco": g.banco or "",
        "valor": float(g.valor or 0),
    }


def gerar_csv(gastos) -> bytes:
    buffer = io.StringIO()
    writer = csv.writer(buffer, delimiter=";")
    writer.writerow(["Data", "Descrição", "Categoria", "Tipo", "Forma de Pagamento", "Banco", "Valor"])
    for g in gastos:
        linha = _linha(g)
        writer.writerow([
            linha["data"], linha["descricao"], linha["categoria"], linha["tipo"],
            linha["forma_pagamento"], linha["banco"],
            f"{linha['valor']:.2f}".replace(".", ","),
        ])
    return buffer.getvalue().encode("utf-8-sig")


def gerar_excel(gastos) -> bytes:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Transações"
    cabecalho = ["Data", "Descrição", "Categoria", "Tipo", "Forma de Pagamento", "Banco", "Valor"]
    ws1.append(cabecalho)
    for cell in ws1[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")

    for g in gastos:
        linha = _linha(g)
        ws1.append([
            linha["data"], linha["descricao"], linha["categoria"], linha["tipo"],
            linha["forma_pagamento"], linha["banco"], linha["valor"],
        ])

    for col in ws1.columns:
        valores = [len(str(c.value)) for c in col if c.value is not None]
        largura = (max(valores) + 2) if valores else 10
        ws1.column_dimensions[col[0].column_letter].width = min(largura, 45)

    ws2 = wb.create_sheet("Por Categoria")
    ws2.append(["Categoria", "Total Gasto"])
    for cell in ws2[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")

    totais_cat: dict[str, float] = {}
    for g in gastos:
        if g.tipo == "saida":
            cat = (g.categoria or "outros").capitalize()
            totais_cat[cat] = totais_cat.get(cat, 0) + float(g.valor or 0)
    for cat, total in sorted(totais_cat.items(), key=lambda x: -x[1]):
        ws2.append([cat, total])

    ws3 = wb.create_sheet("Por Mês")
    ws3.append(["Mês", "Entradas", "Saídas", "Saldo"])
    for cell in ws3[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")

    totais_mes: dict[str, dict[str, float]] = {}
    for g in gastos:
        if not g.data_hora:
            continue
        chave = g.data_hora.strftime("%Y-%m")
        if chave not in totais_mes:
            totais_mes[chave] = {"entrada": 0.0, "saida": 0.0}
        totais_mes[chave][g.tipo] = totais_mes[chave].get(g.tipo, 0) + float(g.valor or 0)
    for chave in sorted(totais_mes.keys()):
        ano, mes = chave.split("-")
        entradas = totais_mes[chave]["entrada"]
        saidas = totais_mes[chave]["saida"]
        ws3.append([f"{mes}/{ano}", entradas, saidas, entradas - saidas])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def gerar_pdf(gastos) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=1.5 * cm, rightMargin=1.5 * cm
    )
    styles = getSampleStyleSheet()
    titulo_style = ParagraphStyle("TituloFinly", parent=styles["Title"], textColor=colors.HexColor("#10b981"))
    elementos = []

    elementos.append(Paragraph("Finly — Relatório de Transações", titulo_style))
    elementos.append(Paragraph(f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles["Normal"]))
    elementos.append(Spacer(1, 16))

    entradas = sum(float(g.valor or 0) for g in gastos if g.tipo == "entrada")
    saidas = sum(float(g.valor or 0) for g in gastos if g.tipo == "saida")
    saldo = entradas - saidas

    resumo = Table([
        ["Entradas", f"R$ {entradas:,.2f}"],
        ["Saídas", f"R$ {saidas:,.2f}"],
        ["Saldo", f"R$ {saldo:,.2f}"],
    ], colWidths=[8 * cm, 8 * cm])
    resumo.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#f8fafc")),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#64748b")),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("TEXTCOLOR", (1, 2), (1, 2), colors.HexColor("#10b981") if saldo >= 0 else colors.HexColor("#ef4444")),
    ]))
    elementos.append(resumo)
    elementos.append(Spacer(1, 20))

    totais_cat: dict[str, float] = {}
    for g in gastos:
        if g.tipo == "saida":
            cat = (g.categoria or "outros").capitalize()
            totais_cat[cat] = totais_cat.get(cat, 0) + float(g.valor or 0)

    if totais_cat:
        elementos.append(Paragraph("Gastos por categoria", styles["Heading2"]))
        itens = sorted(totais_cat.items(), key=lambda x: -x[1])[:10]

        desenho = Drawing(420, 190)
        pizza = Pie()
        pizza.x, pizza.y = 40, 15
        pizza.width, pizza.height = 160, 160
        pizza.data = [v for _, v in itens]
        pizza.labels = [f"{k} (R$ {v:,.0f})" for k, v in itens]
        pizza.slices.strokeWidth = 0.5
        pizza.slices.strokeColor = colors.white
        for i in range(len(itens)):
            pizza.slices[i].fillColor = colors.HexColor(CORES_GRAFICO[i % len(CORES_GRAFICO)])
        desenho.add(pizza)
        elementos.append(desenho)
        elementos.append(Spacer(1, 16))

    elementos.append(Paragraph("Transações", styles["Heading2"]))
    dados = [["Data", "Descrição", "Categoria", "Tipo", "Valor"]]
    for g in gastos[:300]:
        linha = _linha(g)
        dados.append([
            linha["data"], linha["descricao"][:38], linha["categoria"],
            linha["tipo"], f"R$ {linha['valor']:.2f}",
        ])

    tabela = Table(dados, colWidths=[2.2 * cm, 6.3 * cm, 3 * cm, 2.2 * cm, 3 * cm], repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10b981")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#e2e8f0")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8fafc")]),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elementos.append(tabela)

    if len(gastos) > 300:
        elementos.append(Spacer(1, 8))
        elementos.append(Paragraph(
            f"Mostrando as 300 transações mais recentes de {len(gastos)} no total.",
            styles["Italic"]
        ))

    doc.build(elementos)
    return buffer.getvalue()
